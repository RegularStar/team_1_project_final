# certificates/views.py
from collections import defaultdict
from typing import List

from django.db import transaction
from django.db.models import Value, Sum, Avg, Count, Case, When, FloatField, F
from django.db.models.functions import Coalesce
from django.utils.text import slugify
from openpyxl import load_workbook
from rest_framework import filters, permissions, status, viewsets
from rest_framework.pagination import PageNumberPagination
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response

import re


def to_int(value):
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None


class WorksheetUploadMixin:
    def _load_worksheet(self, uploaded_file, request):
        wb = load_workbook(uploaded_file, data_only=True)
        sheet_param = request.query_params.get("sheet")
        if sheet_param:
            try:
                return wb[sheet_param]
            except KeyError:
                raise ValueError(f"시트 '{sheet_param}' 를 찾을 수 없습니다. 시트들: {wb.sheetnames}")
        return wb.active


from .models import (
    Tag,
    Certificate,
    CertificatePhase,
    CertificateStatistics,
    UserTag,
    UserCertificate,
)
from ratings.models import Rating
from .serializers import (
    TagSerializer,
    CertificateSerializer,
    CertificatePhaseSerializer,
    CertificateStatisticsSerializer,
    UserTagSerializer,
    UserCertificateSerializer,
)


class IsAdminOrReadOnly(permissions.BasePermission):
    """GET은 모두, POST/PATCH/DELETE는 관리자만"""
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return True
        return bool(request.user and request.user.is_staff)


# ---- Tag ----
class TagPagination(PageNumberPagination):
    page_size = 30
    page_size_query_param = "page_size"
    max_page_size = 100


class TagViewSet(WorksheetUploadMixin, viewsets.ModelViewSet):
    queryset = Tag.objects.all().order_by("name")
    serializer_class = TagSerializer
    permission_classes = [IsAdminOrReadOnly]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["name"]
    ordering_fields = ["name", "id"]
    pagination_class = TagPagination

    @action(
        detail=False,
        methods=["post"],
        url_path="upload/tags",
        permission_classes=[permissions.IsAdminUser],
        parser_classes=[MultiPartParser, FormParser],
    )
    def upload_tags(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "file 필드를 포함해 업로드하세요."}, status=400)

        try:
            ws = self._load_worksheet(file, request)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=400)

        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        col = {h: i for i, h in enumerate(headers) if h}

        aliases = {
            "id": ["id", "tag_id"],
            "name": ["name", "tag_name"],
        }

        def column_index(key):
            for alias in aliases.get(key, [key]):
                if alias in col:
                    return col[alias]
            return None

        missing = [key for key in ["name"] if column_index(key) is None]
        if missing:
            return Response({"detail": f"누락된 헤더: {', '.join(missing)}"}, status=400)

        def val(row, key):
            idx = column_index(key)
            if idx is None:
                return None
            return row[idx]

        created, updated = 0, 0
        errors: List[str] = []

        with transaction.atomic():
            for r_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row:
                    continue

                raw_name = val(row, "name")
                if raw_name in (None, ""):
                    errors.append(f"{r_index}행: name 누락")
                    continue

                name = str(raw_name).strip()
                if not name:
                    errors.append(f"{r_index}행: name 누락")
                    continue

                tag_id = to_int(val(row, "id"))

                try:
                    if tag_id is not None:
                        obj, is_created = Tag.objects.update_or_create(
                            id=tag_id,
                            defaults={"name": name},
                        )
                    else:
                        obj, is_created = Tag.objects.update_or_create(
                            name=name,
                            defaults={},
                        )
                    if is_created:
                        created += 1
                    else:
                        updated += 1
                except Exception as exc:
                    errors.append(f"{r_index}행: 오류 - {exc}")

        status_code = 200 if not errors else 207
        return Response(
            {
                "created": created,
                "updated": updated,
                "errors": errors,
            },
            status=status_code,
        )


# ---- Certificate ----
class CertificateViewSet(WorksheetUploadMixin, viewsets.ModelViewSet):
    queryset = Certificate.objects.all()
    serializer_class = CertificateSerializer
    permission_classes = [IsAdminOrReadOnly]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["name", "overview", "job_roles", "exam_method", "eligibility", "authority", "type"]
    ordering_fields = ["name", "total_applicants", "avg_difficulty"]

    def get_queryset(self):
        qs = (
            super()
            .get_queryset()
            .prefetch_related("tags")
            .annotate(
                total_applicants=Coalesce(Sum("statistics__applicants"), 0),
                avg_difficulty=Avg("ratings__rating"),
            )
            .order_by("name")
        )

        params = self.request.query_params
        tags_param = params.get("tags")
        if tags_param:
            tag_names = [t.strip() for t in tags_param.split(",") if t.strip()]
            for tag_name in tag_names:
                qs = qs.filter(tags__name__iexact=tag_name)

        type_param = params.get("type")
        if type_param:
            qs = qs.filter(type__iexact=type_param.strip())

        return qs.distinct()

    @action(
        detail=False,
        methods=["post"],
        url_path="upload/certificates",
        permission_classes=[permissions.IsAdminUser],
        parser_classes=[MultiPartParser, FormParser],
    )
    def upload_certificates(self, request):
        """
        XLSX 업로드 (자격증 마스터 전용)
        ✅ 엑셀 맨 앞 열 'id'를 Certificate.pk로 **그대로 사용** (필수)
        ✅ 시트 선택: ?sheet=Certificates (없으면 첫 번째 시트 사용)

        예상 헤더(필수: id, name):
          id, name, overview, job_roles, exam_method, eligibility,
          authority, type, homepage, rating, expected_duration, expected_duration_major, tags

        - 동일 id가 있으면 갱신, 없으면 해당 id로 생성
        - name은 unique라 충돌 시 에러에 기록
        - tags: 'AI,데이터,국가기술' 콤마 구분, 제공 시 set(덮어쓰기), 빈 문자열/None이면 clear()
        """
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "file 필드를 포함해 업로드하세요."}, status=400)

        try:
            ws = self._load_worksheet(file, request)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=400)

        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        col = {h: i for i, h in enumerate(headers) if h}

        aliases = {
            "id": ["id", "cert_id", "certificate_id"],
            "name": ["name"],
            "overview": ["overview", "description"],
            "job_roles": ["job_roles"],
            "exam_method": ["exam_method"],
            "eligibility": ["eligibility"],
            "authority": ["authority"],
            "type": ["type"],
            "homepage": ["homepage"],
            "rating": ["rating"],
            "expected_duration": ["expected_duration"],
            "expected_duration_major": ["expected_duration_major"],
            "tags": ["tags"],
        }

        def column_index(key):
            for alias in aliases.get(key, [key]):
                if alias in col:
                    return col[alias]
            return None

        # id, name은 필수
        missing = [key for key in ["id", "name"] if column_index(key) is None]
        if missing:
            return Response({"detail": f"누락된 헤더: {', '.join(missing)}"}, status=400)

        def val(row, key):
            idx = column_index(key)
            if idx is None:
                return None
            return row[idx]

        created, updated, errors = 0, 0, []

        with transaction.atomic():
            for r_index, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not r:
                    continue

                raw_id = val(r, "id")
                raw_name = val(r, "name")

                # 필수 값 체크
                if raw_id in (None, "") or raw_name in (None, ""):
                    errors.append(f"{r_index}행: id 또는 name 누락")
                    continue

                try:
                    cid = int(raw_id)
                except Exception:
                    errors.append(f"{r_index}행: id가 정수가 아닙니다 → {raw_id}")
                    continue

                defaults = {
                    "name": str(raw_name).strip(),
                    "overview": (str(val(r, "overview")).strip() if val(r, "overview") is not None else ""),
                    "job_roles": (str(val(r, "job_roles")).strip() if val(r, "job_roles") is not None else ""),
                    "exam_method": (str(val(r, "exam_method")).strip() if val(r, "exam_method") is not None else ""),
                    "eligibility": (str(val(r, "eligibility")).strip() if val(r, "eligibility") is not None else ""),
                    "authority": (str(val(r, "authority")).strip() if val(r, "authority") is not None else ""),
                    "type": (str(val(r, "type")).strip() if val(r, "type") is not None else ""),
                    "homepage": (str(val(r, "homepage")).strip() if val(r, "homepage") is not None else ""),
                    "rating": to_int(val(r, "rating")),
                    "expected_duration": to_int(val(r, "expected_duration")),
                    "expected_duration_major": to_int(val(r, "expected_duration_major")),
                }

                try:
                    # 동일 PK가 있으면 갱신, 없으면 명시적 PK로 생성
                    obj = Certificate.objects.filter(pk=cid).first()
                    if obj:
                        for k, v in defaults.items():
                            setattr(obj, k, v)
                        obj.save()
                        updated += 1
                    else:
                        obj = Certificate(id=cid, **defaults)
                        obj.save()
                        created += 1

                    # 태그 동기화(열이 존재할 때만)
                    if "tags" in col:
                        tags_cell = val(r, "tags")
                        if tags_cell in (None, ""):
                            obj.tags.clear()
                        else:
                            tag_names = [t.strip() for t in str(tags_cell).split(",") if t and str(t).strip()]
                            if tag_names:
                                tag_objs = []
                                for tname in tag_names:
                                    tag, _ = Tag.objects.get_or_create(name=tname)
                                    tag_objs.append(tag)
                                obj.tags.set(tag_objs)
                            else:
                                obj.tags.clear()

                except Exception as e:
                    # name unique 충돌 등 모든 예외를 수집
                    errors.append(f"{r_index}행: 오류 - {e}")

        payload = {"created": created, "updated": updated}
        status_code = 200
        if errors:
            payload["errors"] = errors
            status_code = 207
        return Response(payload, status=status_code)

    @action(detail=False, methods=["get"], url_path="rankings")
    def rankings(self, request):
        """Return aggregated ranking data for the home page."""

        try:
            limit = int(request.query_params.get("limit", 10))
        except (TypeError, ValueError):
            limit = 10

        limit = max(1, min(limit, 50))

        certificates = list(
            self.get_queryset()
            .annotate(rating_value=Coalesce("rating", Value(0)))
            .prefetch_related("tags")
        )
        if not certificates:
            return Response(
                {
                    "hot": [],
                    "pass": [],
                    "pass_low": [],
                    "hard_official": [],
                    "easy_official": [],
                    "hard_user": [],
                    "easy_user": [],
                    "difficulty_gap": [],
                    "hell_cards": [],
                }
            )

        cert_map = {cert.id: cert for cert in certificates}

        rating_rows = (
            Rating.objects.filter(certificate_id__in=cert_map)
            .values("certificate_id")
            .annotate(
                average=Avg(
                    Case(
                        When(rating__lte=5, then=F("rating") * Value(2)),
                        default=F("rating"),
                        output_field=FloatField(),
                    )
                ),
                count=Count("id"),
            )
        )
        user_rating_map = {
            row["certificate_id"]: {
                "average": round(float(row["average"]), 1) if row["average"] is not None else None,
                "count": row["count"],
            }
            for row in rating_rows
        }

        stats_by_cert = defaultdict(lambda: defaultdict(dict))

        stats_qs = (
            CertificateStatistics.objects.filter(certificate_id__in=cert_map)
            .values(
                "certificate_id",
                "exam_type",
                "year",
                "registered",
                "applicants",
                "passers",
            )
        )

        def normalize_stage(exam_type):
            if not exam_type:
                return None
            text = str(exam_type).strip()
            digit_match = re.search(r"\d+", text)
            if digit_match:
                try:
                    return int(digit_match.group())
                except Exception:
                    pass

            lowered = text.lower()
            if any(keyword in lowered for keyword in ["필기", "서류", "이론"]):
                return 1
            if any(keyword in lowered for keyword in ["실기", "실습", "작업"]):
                return 2
            if any(keyword in lowered for keyword in ["면접", "구술"]):
                return 3
            if "최종" in lowered:
                return 4
            return None

        def year_key(year):
            if year is None:
                return (-float("inf"), "")
            year_text = str(year).strip()
            digit_match = re.search(r"\d+", year_text)
            if digit_match:
                try:
                    return (int(digit_match.group()), year_text)
                except Exception:
                    pass
            return (0, year_text)

        def format_stage_label(entry, stage_num):
            if stage_num == 10:
                return "전체"
            labels = entry.get("labels") or []
            if labels:
                # Choose the shortest label for readability
                return sorted(labels, key=len)[0]
            return f"{stage_num}차"

        for stat in stats_qs:
            cert_id = stat["certificate_id"]
            stage = normalize_stage(stat.get("exam_type"))
            if stage is None:
                continue
            year = stat.get("year")
            year_key_text = str(year).strip() if year is not None else None
            stage_map = stats_by_cert[cert_id].setdefault(year_key_text, {})
            entry = stage_map.setdefault(
                stage,
                {"registered": 0, "applicants": 0, "passers": 0, "labels": set()},
            )
            for field in ("registered", "applicants", "passers"):
                value = stat.get(field)
                if value is not None:
                    entry[field] += value
            label = stat.get("exam_type")
            if label:
                entry["labels"].add(str(label))

        stage_histories = {}
        for cert_id, year_map in stats_by_cert.items():
            stage_history = defaultdict(list)
            for year_text, stages in year_map.items():
                if not stages:
                    continue
                year_key_value = year_key(year_text)
                for stage_num, entry in stages.items():
                    applicants = to_int(entry.get("applicants"))
                    registered = to_int(entry.get("registered"))
                    passers = to_int(entry.get("passers"))
                    total = applicants if applicants is not None else registered
                    if (total in (None, 0)) and (passers in (None, 0)):
                        continue
                    participant_source = (
                        "applicants"
                        if applicants is not None
                        else ("registered" if registered is not None else None)
                    )
                    pass_rate = None
                    if passers is not None and total not in (None, 0):
                        try:
                            pass_rate = round(passers / total * 100, 1)
                        except ZeroDivisionError:
                            pass_rate = None
                    stage_history[stage_num].append(
                        {
                            "year": year_text,
                            "year_key": year_key_value,
                            "stage": stage_num,
                            "stage_label": format_stage_label(entry, stage_num),
                            "applicants": total,
                            "registered": registered,
                            "raw_applicants": applicants,
                            "passers": passers,
                            "pass_rate": pass_rate,
                            "participant_source": participant_source,
                        }
                    )
            if stage_history:
                for entries in stage_history.values():
                    entries.sort(key=lambda info: info["year_key"], reverse=True)
                stage_histories[cert_id] = stage_history

        def base_item(cert):
            tags = list(cert.tags.order_by("name"))
            tag_names = [tag.name for tag in tags[:10]]
            primary_tag = tag_names[0] if tag_names else None
            rating_info = user_rating_map.get(cert.id, {})
            slug_text = slugify(cert.name) or str(cert.id)
            return {
                "id": cert.id,
                "name": cert.name,
                "slug": slug_text,
                "tag": primary_tag,
                "tags": tag_names,
                "rating": cert.rating,
                "user_difficulty": rating_info.get("average"),
                "user_difficulty_count": rating_info.get("count", 0),
            }

        def format_number(value):
            if value is None:
                return None
            try:
                return f"{int(value):,}"
            except Exception:
                return str(value)

        def stage_applicants_metric(entry):
            count = entry.get("stage_applicants")
            if count is None:
                return None
            source = entry.get("stage_participant_source")
            tooltip_lines = []
            note = entry.get("data_year_note")
            if note:
                tooltip_lines.append(note)
            if source == "registered":
                tooltip_lines.append("응시자 수 집계가 없어 접수 인원을 사용했어요.")
            tooltip = "\n".join(tooltip_lines) if tooltip_lines else None
            return {
                "label": "응시자 수",
                "value": f"{format_number(count)}명",
                "raw": count,
                "tooltip": tooltip,
                "infoButton": bool(tooltip),
            }

        def stage_pass_rate_metric(entry):
            pass_rate = entry.get("stage_pass_rate")
            if pass_rate is None:
                return None
            year = entry.get("recent_year")
            stage_label = entry.get("stage_label")
            base_count = entry.get("stage_applicants")
            passers = entry.get("stage_passers")
            source = entry.get("stage_participant_source")
            tooltip_parts = []
            if year and stage_label:
                tooltip_parts.append(f"{year}년 {stage_label} 합격률")
            elif stage_label:
                tooltip_parts.append(f"{stage_label} 합격률")
            elif year:
                tooltip_parts.append(f"{year}년 합격률")
            if passers is not None and base_count is not None:
                tooltip_parts.append(
                    f"합격자 {format_number(passers)}명 / 응시자 {format_number(base_count)}명"
                )
            if source == "registered":
                tooltip_parts.append("응시자 수 집계가 없어 접수 인원 기준으로 계산한 값이에요.")
            tooltip = "\n".join(tooltip_parts) if tooltip_parts else None
            return {
                "label": "합격률",
                "value": f"{pass_rate:.1f}%",
                "raw": pass_rate,
                "tooltip": tooltip,
            }

        def stage_passers_metric(entry):
            passers = entry.get("stage_passers")
            if passers is None:
                return None
            year = entry.get("recent_year")
            stage_label = entry.get("stage_label")
            tooltip = None
            if year and stage_label:
                tooltip = f"{year}년 {stage_label} 최종 합격자 수"
            elif stage_label:
                tooltip = f"{stage_label} 최종 합격자 수"
            elif year:
                tooltip = f"{year}년 최종 합격자 수"
            return {
                "label": "합격자 수",
                "value": f"{format_number(passers)}명",
                "raw": passers,
                "tooltip": tooltip,
            }

        def build_data_year_label(entry):
            year = entry.get("recent_year")
            if entry.get("is_overall_stage"):
                return f"{year}년 전체 통계" if year else "전체 통계"
            if year:
                return f"{year}년"
            return "최신 공개 통계"

        def build_data_year_note(entry):
            label = entry.get("data_year_label") or build_data_year_label(entry)
            if not label:
                label = "최신 공개 통계"
            return f"몇몇 자격증은 최신자료가 공개되지 않았어요.\n해당 자료는 {label} 기준입니다."

        def format_year_label(year):
            if year in (None, ""):
                return "최신 공개 통계"
            year_text = str(year).strip()
            match = re.search(r"\d{4}", year_text)
            if match:
                return f"{match.group()}년"
            return year_text

        RANK_TOOLTIP_TEXT = (
            "차수별 통계는 최근 공개된 데이터를 기준으로 했어요. 응시자 수 1,000명 이상만 보여줘요."
        )
        DIFFICULTY_RANK_TOOLTIP = "난이도 순위는 SkillBridge 난이도와 사용자 평가를 함께 참고했어요."
        INSIGHT_LIMIT = 8

        def build_stage_records(cert):
            year_data = stats_by_cert.get(cert.id, {})
            if not year_data:
                return []

            base = base_item(cert)
            difficulty_metric = {
                "label": "난이도",
                "value": f"{cert.rating}/10" if cert.rating is not None else None,
                "raw": cert.rating,
                "tooltipKey": "difficulty-scale",
                "tooltip": DIFFICULTY_GUIDE,
            }

            best_by_stage = {}
            for year, stages in year_data.items():
                if not stages:
                    continue
                year_text = year if year not in ("", None) else None
                year_key_value = year_key(year_text)
                for stage_num, entry in stages.items():
                    applicants = to_int(entry.get("applicants"))
                    registered = to_int(entry.get("registered"))
                    passers = to_int(entry.get("passers"))
                    participant_source = "applicants" if applicants is not None else (
                        "registered" if registered is not None else None
                    )
                    total = applicants if applicants is not None else registered
                    if (total is None or total <= 0) and (passers is None or passers <= 0):
                        continue
                    pass_rate = None
                    if total not in (None, 0):
                        pass_rate = round(passers / total * 100, 1) if passers is not None else None
                    stage_label = format_stage_label(entry, stage_num)
                    is_overall = stage_num == 10
                    if is_overall:
                        stage_label = "전체"
                    candidate = best_by_stage.get(stage_num)
                    data = {
                        "stage": stage_num,
                        "stage_label": stage_label,
                        "is_overall_stage": is_overall,
                        "year": year_text,
                        "year_key": year_key_value,
                        "applicants": total,
                        "registered": registered,
                        "passers": passers,
                        "pass_rate": pass_rate,
                        "participant_source": participant_source,
                    }
                    if candidate is None or candidate["year_key"] < year_key_value:
                        best_by_stage[stage_num] = data

            if not best_by_stage:
                return []

            has_specific_stage = any(stage != 10 for stage in best_by_stage)
            if has_specific_stage and 10 in best_by_stage:
                best_by_stage.pop(10)
                if not best_by_stage:
                    return []

            preferred_stage_num = None
            if 1 in best_by_stage:
                preferred_stage_num = 1
            else:
                candidates = list(best_by_stage.values())
                if candidates:
                    preferred = max(
                        candidates,
                        key=lambda info: (
                            1 if not info["is_overall_stage"] else 0,
                            info["applicants"] if info["applicants"] is not None else -1,
                            info["year_key"],
                        ),
                    )
                    preferred_stage_num = preferred["stage"]

            stage_stats = []
            for _, info in sorted(
                best_by_stage.items(),
                key=lambda pair: (pair[0] == 10, pair[0]),
            ):
                stage_stats.append(
                    {
                        "stage": info["stage"],
                        "stage_label": info["stage_label"],
                        "year": info["year"],
                        "pass_rate": info["pass_rate"],
                        "applicants": info["applicants"],
                        "participant_source": info["participant_source"],
                    }
                )

            stage_records = []
            for stage_num, info in sorted(
                best_by_stage.items(),
                key=lambda pair: (pair[0] == 10, pair[0]),
            ):
                record = {
                    **base,
                    "stage": stage_num,
                    "stage_label": info["stage_label"],
                    "is_overall_stage": info["is_overall_stage"],
                    "recent_year": info["year"],
                    "stage_applicants": info["applicants"],
                    "stage_participant_source": info["participant_source"],
                    "stage_passers": info["passers"],
                    "stage_pass_rate": info["pass_rate"],
                    "metric_difficulty": difficulty_metric,
                    "rank_tooltip": RANK_TOOLTIP_TEXT,
                }
                record["data_year_label"] = build_data_year_label(record)
                record["data_year_note"] = build_data_year_note(record)
                record["user_difficulty"] = base.get("user_difficulty")
                record["user_difficulty_count"] = base.get("user_difficulty_count")
                record["is_primary_stage"] = stage_num == preferred_stage_num
                record["is_hell"] = (
                    record.get("rating") is not None
                    and record["rating"] >= 9
                    and (record.get("user_difficulty") or 0) >= 9
                    and (record.get("user_difficulty_count") or 0) > 0
                )
                record["stage_statistics"] = stage_stats
                stage_records.append(record)

            return stage_records

        MIN_STAGE_APPLICANTS = 1000
        stage_payloads = []
        for cert in certificates:
            stage_payloads.extend(build_stage_records(cert))

        eligible_stage_payloads = [
            item
            for item in stage_payloads
            if (item.get("stage_applicants") or 0) >= MIN_STAGE_APPLICANTS
        ]

        primary_stage_payloads = [
            item
            for item in stage_payloads
            if item.get("is_primary_stage") and (item.get("stage_applicants") or 0) >= MIN_STAGE_APPLICANTS
        ]

        def build_difficulty_results(items, *, mode="official", reverse=True):
            def score(entry):
                if mode == "user":
                    return entry.get("user_difficulty")
                return entry.get("rating")

            filtered = []
            for entry in items:
                value = score(entry)
                if value is None:
                    continue
                if mode == "user" and (entry.get("user_difficulty_count") or 0) == 0:
                    continue
                filtered.append(entry)
            if not filtered:
                return []

            sorted_items = sorted(filtered, key=lambda entry: score(entry) or 0, reverse=reverse)
            results = []
            for index, entry in enumerate(sorted_items[:limit], start=1):
                official = entry.get("rating")
                user_value = entry.get("user_difficulty")
                user_count = entry.get("user_difficulty_count") or 0

                official_metric = None
                if official is not None:
                    official_metric = {
                        "label": "공식 난이도",
                        "value": f"{official}/10",
                        "raw": official,
                        "tooltipKey": "difficulty-scale",
                        "tooltip": DIFFICULTY_GUIDE,
                    }

                user_metric_value = f"{user_value:.1f}/10.0" if user_value is not None else "—"
                user_metric_tooltip = None
                if user_count:
                    user_metric_tooltip = f"사용자 {user_count}명이 평가했어요."
                elif user_value is None:
                    user_metric_tooltip = "아직 등록된 사용자 난이도 평가가 없어요."

                if user_value is None and not user_count and mode == "user":
                    continue

                user_metric_display = user_metric_value
                if user_count:
                    user_metric_display = f"{user_metric_value} ({user_count}명 평가)"

                user_metric = {
                    "label": "사용자 난이도",
                    "value": user_metric_display,
                    "raw": user_value,
                    "tooltip": user_metric_tooltip,
                }

                result = {
                    "id": entry["id"],
                    "name": entry["name"],
                    "rank": index,
                    "slug": entry["slug"],
                    "tag": entry.get("tag"),
                    "tags": entry.get("tags"),
                    "rating": entry.get("rating"),
                    "metric": official_metric,
                    "secondary": user_metric,
                    "tertiary": None,
                    "difficulty": None,
                    "rank_tooltip": DIFFICULTY_RANK_TOOLTIP,
                    "data_year": None,
                    "data_year_note": None,
                    "data_year_label": None,
                    "stage": None,
                    "stage_label": None,
                    "is_overall_stage": None,
                    "is_primary_stage": entry.get("is_primary_stage"),
                    "user_difficulty": entry.get("user_difficulty"),
                    "user_difficulty_count": entry.get("user_difficulty_count"),
                    "difference": None,
                    "stage_pass_rate": entry.get("stage_pass_rate"),
                    "stage_passers": entry.get("stage_passers"),
                    "stage_applicants": entry.get("stage_applicants"),
                    "recent_year": entry.get("recent_year"),
                    "is_hell": entry.get("is_hell", False),
                }
                results.append(result)

            return results

        def build_gap_results(items, limit=INSIGHT_LIMIT):
            gap_entries = []
            for entry in items:
                rating = entry.get("rating")
                user = entry.get("user_difficulty")
                user_count = entry.get("user_difficulty_count") or 0
                if rating is None or user is None or user_count == 0:
                    continue
                difference_signed = round(float(user) - float(rating), 1)
                difference_abs = abs(difference_signed)
                if difference_abs <= 0:
                    continue
                gap_entries.append((difference_abs, difference_signed, entry))

            gap_entries.sort(key=lambda item: item[0], reverse=True)
            results = []
            for index, (difference_abs, difference_signed, entry) in enumerate(
                gap_entries[:limit], start=1
            ):
                results.append(
                    {
                        "id": entry["id"],
                        "name": entry["name"],
                        "rank": index,
                        "slug": entry["slug"],
                        "tag": entry.get("tag"),
                        "tags": entry.get("tags"),
                        "rating": entry.get("rating"),
                        "user_difficulty": entry.get("user_difficulty"),
                        "user_difficulty_count": entry.get("user_difficulty_count"),
                        "difference": round(float(difference_abs), 1),
                        "difference_signed": difference_signed,
                        "stage": entry.get("stage"),
                        "stage_label": entry.get("stage_label"),
                        "is_overall_stage": entry.get("is_overall_stage"),
                        "recent_year": entry.get("recent_year"),
                        "stage_applicants": entry.get("stage_applicants"),
                        "stage_pass_rate": entry.get("stage_pass_rate"),
                        "stage_passers": entry.get("stage_passers"),
                        "is_hell": entry.get("is_hell", False),
                    }
                )

            return results

        def build_applicant_change_results(items, limit=INSIGHT_LIMIT):
            results = []
            for entry in items:
                cert_id = entry["id"]
                stage_num = entry.get("stage")
                if stage_num is None:
                    continue
                history_map = stage_histories.get(cert_id) or {}
                stage_history = history_map.get(stage_num) or []
                history_with_applicants = [
                    record for record in stage_history if record.get("applicants") not in (None, 0)
                ]
                if len(history_with_applicants) < 2:
                    continue
                latest = history_with_applicants[0]
                previous = next(
                    (record for record in history_with_applicants[1:] if record.get("applicants") not in (None, 0)),
                    None,
                )
                if not previous:
                    continue
                recent_value = latest.get("applicants") or 0
                previous_value = previous.get("applicants") or 0
                difference = recent_value - previous_value
                if difference == 0:
                    continue
                ratio = None
                if previous_value:
                    ratio = round(difference / previous_value * 100, 1)
                result = {
                    "id": entry["id"],
                    "name": entry["name"],
                    "slug": entry["slug"],
                    "rank": None,
                    "stage": stage_num,
                    "stage_label": latest.get("stage_label") or entry.get("stage_label"),
                    "recent_year": latest.get("year"),
                    "recent_year_label": format_year_label(latest.get("year")),
                    "previous_year": previous.get("year"),
                    "previous_year_label": format_year_label(previous.get("year")),
                    "recent_applicants": recent_value,
                    "previous_applicants": previous_value,
                    "difference": difference,
                    "difference_abs": abs(difference),
                    "difference_ratio": ratio,
                    "participant_source": latest.get("participant_source"),
                    "recent_pass_rate": latest.get("pass_rate"),
                    "previous_pass_rate": previous.get("pass_rate"),
                    "tag": entry.get("tag"),
                    "tags": entry.get("tags"),
                    "rating": entry.get("rating"),
                    "user_difficulty": entry.get("user_difficulty"),
                    "user_difficulty_count": entry.get("user_difficulty_count"),
                }
                results.append(result)
            if not results:
                return []
            results.sort(
                key=lambda item: (
                    item["difference_abs"],
                    abs(item.get("difference_ratio") or 0),
                    item.get("recent_applicants") or 0,
                ),
                reverse=True,
            )
            limited = []
            for index, entry in enumerate(results[:limit], start=1):
                entry["rank"] = index
                limited.append(entry)
            return limited

        def build_stage_pass_gap_results(items, limit=INSIGHT_LIMIT):
            results = []
            for entry in items:
                cert_id = entry["id"]
                history_map = stage_histories.get(cert_id) or {}
                stage1_history = history_map.get(1) or []
                stage2_history = history_map.get(2) or []
                if not stage1_history or not stage2_history:
                    continue
                stage1_by_year = {
                    record["year"]: record for record in stage1_history if record.get("pass_rate") is not None
                }
                if not stage1_by_year:
                    continue
                best_pair = None
                for second in stage2_history:
                    pass_rate_two = second.get("pass_rate")
                    if pass_rate_two is None:
                        continue
                    year_key_value = second.get("year_key")
                    matching = stage1_by_year.get(second.get("year"))
                    if not matching or matching.get("pass_rate") is None:
                        continue
                    if best_pair is None or year_key_value > best_pair[0]:
                        best_pair = (year_key_value, matching, second)
                if not best_pair:
                    continue
                _, stage1_entry, stage2_entry = best_pair
                diff_value = abs(stage1_entry["pass_rate"] - stage2_entry["pass_rate"])
                if diff_value <= 0:
                    continue
                result = {
                    "id": entry["id"],
                    "name": entry["name"],
                    "slug": entry["slug"],
                    "rank": None,
                    "stage": entry.get("stage"),
                    "stage_label": entry.get("stage_label"),
                    "year": stage1_entry.get("year") or stage2_entry.get("year"),
                    "year_label": format_year_label(stage1_entry.get("year") or stage2_entry.get("year")),
                    "stage1_label": stage1_entry.get("stage_label") or "1차",
                    "stage2_label": stage2_entry.get("stage_label") or "2차",
                    "stage1_pass_rate": stage1_entry.get("pass_rate"),
                    "stage2_pass_rate": stage2_entry.get("pass_rate"),
                    "stage1_applicants": stage1_entry.get("applicants"),
                    "stage2_applicants": stage2_entry.get("applicants"),
                    "stage1_participant_source": stage1_entry.get("participant_source"),
                    "stage2_participant_source": stage2_entry.get("participant_source"),
                    "difference": round(diff_value, 1),
                    "difference_signed": round(
                        stage1_entry.get("pass_rate") - stage2_entry.get("pass_rate"), 1
                    ),
                    "tag": entry.get("tag"),
                    "tags": entry.get("tags"),
                    "rating": entry.get("rating"),
                    "user_difficulty": entry.get("user_difficulty"),
                    "user_difficulty_count": entry.get("user_difficulty_count"),
                }
                results.append(result)
            if not results:
                return []
            results.sort(key=lambda item: item["difference"], reverse=True)
            limited = []
            for index, entry in enumerate(results[:limit], start=1):
                entry["rank"] = index
                limited.append(entry)
            return limited

        def build_hell_results(items, limit=5):
            hell_entries = [
                entry
                for entry in items
                if entry.get("is_hell")
            ]
            if not hell_entries:
                return []

            hell_entries.sort(
                key=lambda entry: (
                    entry.get("user_difficulty") or 0,
                    entry.get("rating") or 0,
                ),
                reverse=True,
            )
            results = []
            for entry in hell_entries[:limit]:
                results.append(
                    {
                        "id": entry["id"],
                        "name": entry["name"],
                        "slug": entry["slug"],
                        "tag": entry.get("tag"),
                        "tags": entry.get("tags"),
                        "rating": entry.get("rating"),
                        "user_difficulty": entry.get("user_difficulty"),
                        "user_difficulty_count": entry.get("user_difficulty_count"),
                        "stage": entry.get("stage"),
                        "stage_label": entry.get("stage_label"),
                        "is_overall_stage": entry.get("is_overall_stage"),
                        "recent_year": entry.get("recent_year"),
                        "stage_applicants": entry.get("stage_applicants"),
                        "stage_pass_rate": entry.get("stage_pass_rate"),
                        "stage_passers": entry.get("stage_passers"),
                        "stage_statistics": entry.get("stage_statistics"),
                        "is_hell": True,
                    }
                )

            return results

        def sort_and_build(items, key_func, metric_selector, secondary_selector=None, tertiary_selector=None):
            sorted_items = [item for item in items if key_func(item) is not None]
            sorted_items.sort(key=key_func, reverse=True)
            results = []
            for index, entry in enumerate(sorted_items[:limit], start=1):
                metric = metric_selector(entry)
                secondary = secondary_selector(entry) if secondary_selector else None
                tertiary = tertiary_selector(entry) if tertiary_selector else None
                results.append(
                    {
                        "id": entry["id"],
                        "name": entry["name"],
                        "rank": index,
                        "slug": entry["slug"],
                        "tag": entry.get("tag"),
                        "tags": entry.get("tags"),
                        "rating": entry.get("rating"),
                        "metric": metric,
                        "secondary": secondary,
                        "tertiary": tertiary,
                        "difficulty": entry.get("metric_difficulty"),
                        "rank_tooltip": entry.get("rank_tooltip"),
                        "data_year": entry.get("recent_year"),
                        "data_year_note": entry.get("data_year_note"),
                        "data_year_label": entry.get("data_year_label"),
                        "stage": entry.get("stage"),
                        "stage_label": entry.get("stage_label"),
                        "is_overall_stage": entry.get("is_overall_stage"),
                        "is_primary_stage": entry.get("is_primary_stage"),
                        "user_difficulty": entry.get("user_difficulty"),
                        "user_difficulty_count": entry.get("user_difficulty_count"),
                        "stage_pass_rate": entry.get("stage_pass_rate"),
                        "stage_passers": entry.get("stage_passers"),
                        "stage_applicants": entry.get("stage_applicants"),
                        "is_hell": entry.get("is_hell", False),
                    }
                )
            return results

        hot_items = sort_and_build(
            eligible_stage_payloads,
            key_func=lambda item: item.get("stage_applicants"),
            metric_selector=stage_applicants_metric,
            secondary_selector=stage_pass_rate_metric,
            tertiary_selector=stage_passers_metric,
        )

        pass_items = sort_and_build(
            eligible_stage_payloads,
            key_func=lambda item: item.get("stage_pass_rate"),
            metric_selector=stage_applicants_metric,
            secondary_selector=stage_pass_rate_metric,
            tertiary_selector=stage_passers_metric,
        )

        hard_official = build_difficulty_results(primary_stage_payloads, mode="official", reverse=True)

        easy_official = build_difficulty_results(primary_stage_payloads, mode="official", reverse=False)

        hard_user = build_difficulty_results(primary_stage_payloads, mode="user", reverse=True)

        easy_user = build_difficulty_results(primary_stage_payloads, mode="user", reverse=False)

        difficulty_gap = build_gap_results(primary_stage_payloads)

        hell_cards = build_hell_results(primary_stage_payloads)
        for card in hell_cards:
            card["badge_label"] = "지옥의 자격증"
            card["badge_variant"] = "hell"

        applicant_surge = build_applicant_change_results(eligible_stage_payloads)

        stage_pass_gap = build_stage_pass_gap_results(primary_stage_payloads)

        MAJOR_PROFESSIONALS = [
            "변호사",
            "공인회계사",
            "변리사",
            "공인노무사",
            "세무사",
            "법무사",
            "감정평가사",
            "관세사",
        ]
        def find_badge_entry(target_name: str):
            normalized = (target_name or "").strip()
            if not normalized:
                return None

            def match_pool(pool):
                for item in pool:
                    item_name = (item.get("name") or "").strip()
                    if not item_name:
                        continue
                    if item_name == normalized:
                        return item
                for item in pool:
                    item_name = (item.get("name") or "").strip()
                    if not item_name:
                        continue
                    if normalized in item_name:
                        return item
                return None

            entry = match_pool(primary_stage_payloads)
            if entry:
                return entry
            return match_pool(stage_payloads)

        def serialize_badge_entry(entry, *, badge_label, badge_variant):
            return {
                "id": entry["id"],
                "name": entry["name"],
                "slug": entry["slug"],
                "tag": entry.get("tag"),
                "tags": entry.get("tags"),
                "rating": entry.get("rating"),
                "user_difficulty": entry.get("user_difficulty"),
                "user_difficulty_count": entry.get("user_difficulty_count"),
                "stage": entry.get("stage"),
                "stage_label": entry.get("stage_label"),
                "is_overall_stage": entry.get("is_overall_stage"),
                "recent_year": entry.get("recent_year"),
                "stage_applicants": entry.get("stage_applicants"),
                "stage_participant_source": entry.get("stage_participant_source"),
                "stage_pass_rate": entry.get("stage_pass_rate"),
                "stage_passers": entry.get("stage_passers"),
                "stage_statistics": entry.get("stage_statistics"),
                "badge_label": badge_label,
                "badge_variant": badge_variant,
            }

        professional_badges = []
        seen_badge_ids = set()
        for name in MAJOR_PROFESSIONALS:
            entry = find_badge_entry(name)
            if not entry:
                continue
            entry_id = entry.get("id")
            if entry_id in seen_badge_ids:
                continue
            seen_badge_ids.add(entry_id)
            professional_badges.append(
                serialize_badge_entry(entry, badge_label="8대 전문직", badge_variant="elite")
            )

        badge_groups = [
            {
                "key": "hell",
                "title": "지옥의 자격증",
                "variant": "hell",
                "ribbon": "지옥의 자격증",
                "items": hell_cards,
            },
            {
                "key": "elite",
                "title": "8대 전문직",
                "variant": "elite",
                "ribbon": "8대 전문직",
                "items": professional_badges,
            },
        ]

        insight_groups = [
            {
                "key": "difficulty_gap",
                "title": "체감 난이도",
                "subtitle": "공식 난이도와 사용자 난이도 차이가 큰 자격증",
                "items": difficulty_gap,
            },
            {
                "key": "applicant_surge",
                "title": "응시자 수 변화",
                "subtitle": "최근 2개 년도 기준 응시자 수 변동이 큰 자격증",
                "items": applicant_surge,
            },
            {
                "key": "stage_pass_gap",
                "title": "시험단계별 합격률 격차",
                "subtitle": "1차와 2차 합격률 온도차가 큰 자격증",
                "items": stage_pass_gap,
            },
        ]

        pass_low_items = sort_and_build(
            eligible_stage_payloads,
            key_func=lambda item: (
                -item.get("stage_pass_rate")
                if item.get("stage_pass_rate") is not None
                else None
            ),
            metric_selector=stage_applicants_metric,
            secondary_selector=stage_pass_rate_metric,
            tertiary_selector=stage_passers_metric,
        )

        data = {
            "hot": hot_items,
            "pass": pass_items,
            "pass_low": pass_low_items,
            "hard_official": hard_official,
            "easy_official": easy_official,
            "hard_user": hard_user,
            "easy_user": easy_user,
            "difficulty_gap": difficulty_gap,
            "hell_cards": hell_cards,
            "applicant_surge": applicant_surge,
            "stage_pass_gap": stage_pass_gap,
            "insight_groups": insight_groups,
            "badge_groups": badge_groups,
        }

        return Response(data)

    @action(
        detail=False,
        methods=["post"],
        url_path="upload/certificate-tags",
        permission_classes=[permissions.IsAdminUser],
        parser_classes=[MultiPartParser, FormParser],
    )
    def upload_certificate_tags(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "file 필드를 포함해 업로드하세요."}, status=400)

        try:
            ws = self._load_worksheet(file, request)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=400)

        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        col = {h: i for i, h in enumerate(headers) if h}

        aliases = {
            "certificate_id": ["certificate_id", "cert_id", "certificate"],
            "certificate_name": ["certificate_name", "cert_name"],
            "tags": ["tags", "tag_names"],
            "tag_ids": ["tag_ids"],
        }

        def column_index(key):
            for alias in aliases.get(key, [key]):
                if alias in col:
                    return col[alias]
            return None

        if column_index("certificate_id") is None and column_index("certificate_name") is None:
            return Response({"detail": "certificate_id 또는 certificate_name 열이 필요합니다."}, status=400)

        if column_index("tags") is None and column_index("tag_ids") is None:
            return Response({"detail": "tags 또는 tag_ids 열 중 하나는 포함되어야 합니다."}, status=400)

        def val(row, key):
            idx = column_index(key)
            if idx is None:
                return None
            return row[idx]

        updated, cleared = 0, 0
        created_tags = 0
        errors: List[str] = []

        with transaction.atomic():
            for r_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row:
                    continue

                cert = None
                cert_id = to_int(val(row, "certificate_id"))
                cert_name_raw = val(row, "certificate_name")
                if cert_id is not None:
                    cert = Certificate.objects.filter(id=cert_id).first()
                if cert is None and cert_name_raw not in (None, ""):
                    cert = Certificate.objects.filter(name=str(cert_name_raw).strip()).first()

                if cert is None:
                    errors.append(f"{r_index}행: 자격증을 찾을 수 없습니다.")
                    continue

                tag_objs: List[Tag] = []
                tag_names_raw = val(row, "tags")
                tag_ids_raw = val(row, "tag_ids")

                def split_values(raw):
                    if raw in (None, ""):
                        return []
                    items = re.split(r"[,\n;/]", str(raw))
                    return [item.strip() for item in items if item and item.strip()]

                tag_names = split_values(tag_names_raw)
                tag_ids = split_values(tag_ids_raw)

                tag_id_objs: List[Tag] = []
                for raw_id in tag_ids:
                    tag_obj = None
                    tag_pk = to_int(raw_id)
                    if tag_pk is not None:
                        tag_obj = Tag.objects.filter(id=tag_pk).first()
                    if not tag_obj:
                        errors.append(f"{r_index}행: tag_id '{raw_id}' 를 찾을 수 없습니다.")
                        tag_id_objs = []
                        break
                    tag_id_objs.append(tag_obj)
                if errors and (not tag_id_objs and tag_ids):
                    continue

                for name in tag_names:
                    tag_obj, is_created = Tag.objects.get_or_create(name=name)
                    if is_created:
                        created_tags += 1
                    tag_objs.append(tag_obj)

                tag_objs.extend(tag_id_objs)

                if not tag_objs:
                    cert.tags.clear()
                    cleared += 1
                    continue

                unique_tag_ids = []
                seen_ids = set()
                for tag_obj in tag_objs:
                    if tag_obj.id not in seen_ids:
                        seen_ids.add(tag_obj.id)
                        unique_tag_ids.append(tag_obj)

                cert.tags.set(unique_tag_ids)
                updated += 1

        status_code = 200 if not errors else 207
        return Response(
            {
                "updated_certificates": updated,
                "cleared_certificates": cleared,
                "created_tags": created_tags,
                "errors": errors,
            },
            status=status_code,
        )


# ---- Phase ----
class CertificatePhaseViewSet(WorksheetUploadMixin, viewsets.ModelViewSet):
    queryset = CertificatePhase.objects.select_related("certificate").all()
    serializer_class = CertificatePhaseSerializer
    permission_classes = [IsAdminOrReadOnly]

    @action(
        detail=False,
        methods=["post"],
        url_path="upload/phases",
        permission_classes=[permissions.IsAdminUser],
        parser_classes=[MultiPartParser, FormParser],
    )
    def upload_phases(self, request):
        """자격증 단계 일괄 업로드."""
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "file 필드를 포함해 업로드하세요."}, status=400)

        try:
            ws = self._load_worksheet(file, request)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=400)

        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        col = {h: i for i, h in enumerate(headers) if h}

        aliases = {
            "id": ["id"],
            "certificate_id": ["certificate_id", "cert_id"],
            "certificate_name": ["certificate_name", "cert_name"],
            "phase_name": ["phase_name"],
            "phase_type": ["phase_type", "phase_category"],
        }

        def column_index(key):
            for alias in aliases.get(key, [key]):
                if alias in col:
                    return col[alias]
            return None

        def val(row, key):
            idx = column_index(key)
            if idx is None:
                return None
            return row[idx]

        created, updated = 0, 0
        with transaction.atomic():
            for r in ws.iter_rows(min_row=2, values_only=True):
                if not r:
                    continue

                phase_id = to_int(val(r, "id"))

                cert = None
                cert_pk = to_int(val(r, "certificate_id"))
                if cert_pk is not None:
                    cert = Certificate.objects.filter(id=cert_pk).first()
                if cert is None:
                    raw_cert_name = val(r, "certificate_name")
                    if raw_cert_name not in (None, ""):
                        cert = Certificate.objects.filter(name=str(raw_cert_name).strip()).first()
                if cert is None:
                    continue

                raw_phase_name = val(r, "phase_name")
                phase_name = str(raw_phase_name).strip() if raw_phase_name not in (None, "") else None
                if not phase_name:
                    continue

                raw_phase_type = val(r, "phase_type")
                phase_type = str(raw_phase_type).strip() if raw_phase_type not in (None, "") else ""

                if phase_id is not None:
                    defaults = {
                        "certificate": cert,
                        "phase_name": phase_name,
                        "phase_type": phase_type,
                    }
                    _, is_created = CertificatePhase.objects.update_or_create(
                        id=phase_id,
                        defaults=defaults,
                    )
                else:
                    _, is_created = CertificatePhase.objects.update_or_create(
                        certificate=cert,
                        phase_name=phase_name,
                        phase_type=phase_type,
                        defaults={},
                    )

                created += int(is_created)
                updated += int(not is_created)

        return Response({"created": created, "updated": updated}, status=200)

# ---- Statistics ----
class CertificateStatisticsViewSet(WorksheetUploadMixin, viewsets.ModelViewSet):
    queryset = CertificateStatistics.objects.select_related("certificate").all()
    serializer_class = CertificateStatisticsSerializer
    permission_classes = [IsAdminOrReadOnly]

    @action(
        detail=False,
        methods=["post"],
        url_path="upload/statistics",
        permission_classes=[permissions.IsAdminUser],
        parser_classes=[MultiPartParser, FormParser],
    )
    def upload_statistics(self, request):
        """자격증 통계 일괄 업로드."""
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "file 필드를 포함해 업로드하세요."}, status=400)

        try:
            ws = self._load_worksheet(file, request)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=400)

        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        col = {h: i for i, h in enumerate(headers) if h}

        aliases = {
            "id": ["id", "stat_id", "cert_stats_id"],
            "certificate_id": ["certificate_id", "cert_id"],
            "certificate_name": ["certificate_name", "cert_name"],
            "exam_type": ["exam_type"],
            "year": ["year"],
            "session": ["session"],
            "registered": ["registered", "registerd"],
            "applicants": ["applicants"],
            "passers": ["passers"],
            "pass_rate": ["pass_rate"],
        }

        def column_index(key):
            for alias in aliases.get(key, [key]):
                if alias in col:
                    return col[alias]
            return None

        def val(row, key):
            idx = column_index(key)
            if idx is None:
                return None
            return row[idx]

        def normalize_rate(x):
            if x in (None, ""):
                return None
            try:
                f = float(x)
                if 0 <= f <= 1:
                    return round(f * 100, 1)
                if 0 <= f <= 100:
                    return round(f, 1)
            except Exception:
                pass
            return None

        created, updated = 0, 0
        with transaction.atomic():
            for r in ws.iter_rows(min_row=2, values_only=True):
                if not r:
                    continue

                stat_id = to_int(val(r, "id"))

                cert = None
                cert_pk = to_int(val(r, "certificate_id"))
                if cert_pk is not None:
                    cert = Certificate.objects.filter(id=cert_pk).first()
                if cert is None:
                    raw_cert_name = val(r, "certificate_name")
                    if raw_cert_name not in (None, ""):
                        cert = Certificate.objects.filter(name=str(raw_cert_name).strip()).first()
                if cert is None:
                    continue

                exam_type = str(val(r, "exam_type") or "").strip() or "필기"
                year_raw = val(r, "year")
                year = str(year_raw).strip() if year_raw not in (None, "") else None
                if not year:
                    continue

                session = to_int(val(r, "session"))
                registered = to_int(val(r, "registered"))
                applicants = to_int(val(r, "applicants"))
                passers = to_int(val(r, "passers"))
                pass_rate = normalize_rate(val(r, "pass_rate"))

                if stat_id is not None:
                    defaults = {
                        "certificate": cert,
                        "exam_type": exam_type,
                        "year": year,
                        "session": session,
                        "registered": registered,
                        "applicants": applicants,
                        "passers": passers,
                        "pass_rate": pass_rate,
                    }
                    _, is_created = CertificateStatistics.objects.update_or_create(
                        id=stat_id,
                        defaults=defaults,
                    )
                else:
                    _, is_created = CertificateStatistics.objects.update_or_create(
                        certificate=cert,
                        exam_type=exam_type,
                        year=year,
                        session=session,
                        defaults={
                            "registered": registered,
                            "applicants": applicants,
                            "passers": passers,
                            "pass_rate": pass_rate,
                        },
                    )

                created += int(is_created)
                updated += int(not is_created)

        return Response({"created": created, "updated": updated}, status=200)


# ---- UserTag (내 태그만 관리) ----
class UserTagViewSet(viewsets.ModelViewSet):
    serializer_class = UserTagSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return UserTag.objects.select_related("tag").filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


# ---- UserCertificate (내 취득 자격증 관리) ----
class UserCertificateViewSet(viewsets.ModelViewSet):
    serializer_class = UserCertificateSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return (
            UserCertificate.objects.select_related("certificate", "reviewed_by")
            .filter(user=self.request.user)
            .order_by("-created_at")
        )

    def perform_create(self, serializer):
        serializer.save(
            user=self.request.user,
            status=UserCertificate.STATUS_PENDING,
            review_note="",
            reviewed_by=None,
            reviewed_at=None,
        )

    def perform_update(self, serializer):
        # 사용자가 정보를 수정하면 다시 심사 상태로 전환한다.
        instance = serializer.save(
            status=UserCertificate.STATUS_PENDING,
            review_note="",
            reviewed_by=None,
            reviewed_at=None,
        )
        return instance

DIFFICULTY_GUIDE = (
    "난이도 안내\n"
    "1: 아주 쉬움. 기초 개념 위주라 단기간 준비로 누구나 합격 가능한 수준.\n"
    "2: 쉬움. 기본 지식이 있으면 무난히 도전할 수 있는 입문 수준.\n"
    "3: 보통. 일정한 학습이 필요하지만 꾸준히 준비하면 충분히 합격 가능한 수준.\n"
    "4: 다소 어려움. 이론과 실무를 균형 있게 요구하며, 준비 기간이 다소 긴 수준.\n"
    "5: 중상 난이도. 전공지식과 응용력이 필요해 체계적 학습이 요구되는 수준.\n"
    "6: 어려움. 합격률이 낮고 심화 학습이 필요해 전공자도 부담되는 수준.\n"
    "7: 매우 어려움. 방대한 범위와 높은 난이도로 전공자도 장기간 학습이 필수인 수준.\n"
    "8: 극히 어려움. 전문성·응용력·실무 경험이 모두 요구되는 최상위권 자격 수준.\n"
    "9: 최상 난이도. 전문지식과 실무를 총망라하며, 합격자가 극소수에 불과한 수준.\n"
    "10: 극한 난이도. 수년간 전념해도 합격을 장담할 수 없는, 최고 난도의 자격 수준."
)
