#!/usr/bin/env python3
"""
Utility script to build RAG-friendly documents from the data spreadsheet.

Example usage:
    python scripts/build_rag_documents.py \
        --input data/data.xlsx \
        --output data/rag/documents.jsonl
"""

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook


@dataclass
class RatingInfo:
    score: str
    description: Optional[str]


@dataclass
class CertificateInfo:
    cert_id: str
    name: str
    overview: Optional[str]
    job_roles: Optional[str]
    exam_method: Optional[str]
    eligibility: Optional[str]
    rating: Optional[str]
    expected_duration: Optional[str]
    expected_duration_major: Optional[str]
    authority: Optional[str]
    cert_type: Optional[str]
    homepage: Optional[str]


@dataclass
class StatisticEntry:
    cert_id: str
    stat_id: Optional[str]
    exam_type: Optional[str]
    stage: Optional[int]
    year: Optional[str]
    session: Optional[str]
    registered: Optional[int]
    applicants: Optional[int]
    passers: Optional[int]
    pass_rate: Optional[float]


def load_table_rows(worksheet) -> List[Dict[str, Any]]:
    headers: List[str] = []
    for cell in worksheet[1]:
        value = cell.value
        headers.append("" if value is None else str(value).strip())

    rows: List[Dict[str, Any]] = []
    for raw in worksheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in raw):
            continue
        row: Dict[str, Any] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            row[header] = raw[index] if index < len(raw) else None
        rows.append(row)
    return rows


def to_int(value: Any) -> Optional[int]:
    if value in (None, "", "null"):
        return None
    try:
        if isinstance(value, float):
            if value != value:  # NaN
                return None
            return int(round(value))
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def to_float(value: Any) -> Optional[float]:
    if value in (None, "", "null"):
        return None
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def normalize_text(value: Any) -> Optional[str]:
    if value in (None, ""):
        return None
    text = str(value).strip()
    return text or None


def normalize_stage(exam_type: Optional[str]) -> Optional[int]:
    if exam_type is None:
        return None
    text = str(exam_type).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError):
        pass

    digit_match = re.search(r"\d+", text)
    if digit_match:
        try:
            return int(digit_match.group())
        except (TypeError, ValueError):
            pass

    lowered = text.lower()
    if any(keyword in lowered for keyword in ["필기", "서류", "이론"]):
        return 1
    if any(keyword in lowered for keyword in ["실기", "실습", "작업"]):
        return 2
    if any(keyword in lowered for keyword in ["면접", "구술"]):
        return 3
    if "최종" in lowered:
        return 4
    if "전체" in lowered:
        return 10
    return None


def format_stage_label(stage: Optional[int], exam_type: Optional[str]) -> str:
    if stage == 10:
        return "전체"
    exam_text = normalize_text(exam_type)
    if exam_text:
        return exam_text
    if stage is not None:
        return f"{stage}차"
    return "차수 미상"


def format_year(text: Optional[str]) -> str:
    if text in (None, "", "None"):
        return "연도 미상"
    clean = str(text).strip()
    return f"{clean}년" if clean.isdigit() else clean


def format_number(value: Optional[int]) -> Optional[str]:
    if value is None:
        return None
    try:
        return f"{int(value):,d}"
    except (TypeError, ValueError):
        return None


def format_percentage(value: Optional[float]) -> Optional[str]:
    if value is None:
        return None
    return f"{value:.1f}%"


def build_rating_map(rows: Iterable[Dict[str, Any]]) -> Dict[str, RatingInfo]:
    mapping: Dict[str, RatingInfo] = {}
    for row in rows:
        score = normalize_text(row.get("rating"))
        if not score:
            continue
        description = normalize_text(row.get("description"))
        mapping[score] = RatingInfo(score=score, description=description)
    return mapping


def build_certificate_rows(rows: Iterable[Dict[str, Any]]) -> Dict[str, CertificateInfo]:
    certificates: Dict[str, CertificateInfo] = {}
    for row in rows:
        cert_id = normalize_text(row.get("id"))
        name = normalize_text(row.get("name"))
        if not cert_id or not name:
            continue
        certificates[cert_id] = CertificateInfo(
            cert_id=cert_id,
            name=name,
            overview=normalize_text(row.get("overview")),
            job_roles=normalize_text(row.get("job_roles")),
            exam_method=normalize_text(row.get("exam_method")),
            eligibility=normalize_text(row.get("eligibility")),
            rating=normalize_text(row.get("rating")),
            expected_duration=normalize_text(row.get("expected_duration")),
            expected_duration_major=normalize_text(row.get("expected_duration_major")),
            authority=normalize_text(row.get("authority")),
            cert_type=normalize_text(row.get("type")),
            homepage=normalize_text(row.get("homepage")),
        )
    return certificates


def build_statistics(rows: Iterable[Dict[str, Any]]) -> List[StatisticEntry]:
    stats: List[StatisticEntry] = []
    for row in rows:
        cert_id = normalize_text(row.get("cert_id")) or normalize_text(row.get("certificate_id"))
        if not cert_id:
            continue
        exam_type = normalize_text(row.get("exam_type"))
        stage = normalize_stage(exam_type)
        year = normalize_text(row.get("year"))
        session = normalize_text(row.get("session"))
        registered = to_int(row.get("registered") or row.get("registerd"))
        applicants = to_int(row.get("applicants"))
        passers = to_int(row.get("passers"))
        pass_rate = to_float(row.get("pass_rate"))

        base_total = applicants if applicants not in (None, 0) else registered
        calculated_rate: Optional[float] = None
        if passers not in (None, 0) and base_total not in (None, 0):
            try:
                calculated_rate = round(passers / base_total * 100, 1)
            except ZeroDivisionError:
                calculated_rate = None

        if pass_rate is None:
            pass_rate = calculated_rate
        else:
            pass_rate = round(pass_rate, 1)
            if pass_rate <= 1 and calculated_rate and calculated_rate > 1:
                # 워크북에 0~1 범위 비율로 저장된 값 보정
                pass_rate = calculated_rate

        stats.append(
            StatisticEntry(
                cert_id=cert_id,
                stat_id=normalize_text(row.get("id")),
                exam_type=exam_type,
                stage=stage,
                year=year,
                session=session,
                registered=registered,
                applicants=applicants,
                passers=passers,
                pass_rate=pass_rate,
            )
        )
    return stats


def build_profile_document(cert: CertificateInfo, rating_map: Dict[str, RatingInfo]) -> Dict[str, Any]:
    lines: List[str] = []
    lines.append(f"{cert.name} 자격증 정보")
    lines.append("")

    rating_info = rating_map.get(cert.rating or "")
    if cert.rating:
        rating_text = f"난이도 {cert.rating}"
        if rating_info and rating_info.description:
            rating_text += f" - {rating_info.description}"
        lines.append(rating_text)
    elif rating_info and rating_info.description:
        lines.append(rating_info.description)

    if cert.overview:
        lines.append(f"개요: {cert.overview}")
    if cert.job_roles:
        lines.append(f"활용 직무: {cert.job_roles}")
    if cert.exam_method:
        lines.append(f"시험 방식: {cert.exam_method}")
    if cert.eligibility:
        lines.append(f"응시 자격: {cert.eligibility}")
    if cert.expected_duration:
        lines.append(f"예상 취득 기간(전공자): {cert.expected_duration}")
    if cert.expected_duration_major:
        lines.append(f"예상 취득 기간(비전공자): {cert.expected_duration_major}")
    if cert.authority:
        lines.append(f"시행 기관: {cert.authority}")
    if cert.cert_type:
        lines.append(f"자격증 유형: {cert.cert_type}")
    if cert.homepage:
        lines.append(f"공식 홈페이지: {cert.homepage}")

    if len(lines) == 2:
        lines.append("추가 정보가 제공되지 않았습니다.")

    text = "\n".join(lines).strip()
    return {
        "id": f"certificate_profile:{cert.cert_id}",
        "certificate_id": cert.cert_id,
        "type": "certificate_profile",
        "name": cert.name,
        "text": text,
    }


def build_statistics_documents(
    cert: CertificateInfo,
    entries: Iterable[StatisticEntry],
) -> List[Dict[str, Any]]:
    grouped: Dict[Optional[str], List[StatisticEntry]] = defaultdict(list)
    for entry in entries:
        grouped[entry.year].append(entry)

    documents: List[Dict[str, Any]] = []
    for year, year_entries in sorted(grouped.items(), key=lambda item: (item[0] or "",), reverse=True):
        if not year_entries:
            continue
        lines: List[str] = []
        lines.append(f"{cert.name} 통계 요약 - {format_year(year)}")
        lines.append("")
        year_entries_sorted = sorted(
            year_entries,
            key=lambda x: (x.stage if x.stage is not None else 99, normalize_text(x.exam_type) or ""),
        )

        for entry in year_entries_sorted:
            stage_label = format_stage_label(entry.stage, entry.exam_type)
            parts: List[str] = []
            applicants = entry.applicants
            registered = entry.registered if entry.registered not in (None, 0) else None
            passers = entry.passers
            pass_rate = entry.pass_rate

            if applicants is not None:
                parts.append(f"응시자 {format_number(applicants)}명")
            elif registered is not None:
                parts.append(f"접수자 {format_number(registered)}명")
            if passers is not None:
                parts.append(f"합격자 {format_number(passers)}명")

            if pass_rate is None and passers is not None:
                base = applicants if applicants not in (None, 0) else registered
                if base not in (None, 0):
                    pass_rate = round(passers / base * 100, 1)
                elif base == 0:
                    pass_rate = 0.0

            rate_text = format_percentage(pass_rate)
            if rate_text:
                parts.append(f"합격률 {rate_text}")

            if entry.session:
                parts.append(f"시행 {entry.session}회")

            if not parts:
                parts.append("수치 정보 없음")

            lines.append(f"- {stage_label}: {', '.join(parts)}")

        documents.append(
            {
                "id": f"certificate_stats:{cert.cert_id}:{year or 'unknown'}",
                "certificate_id": cert.cert_id,
                "type": "certificate_statistics",
                "name": cert.name,
                "year": year,
                "text": "\n".join(lines).strip(),
            }
        )

    return documents


def generate_documents(input_path: Path) -> List[Dict[str, Any]]:
    workbook = load_workbook(input_path, data_only=True)
    rating_rows = load_table_rows(workbook["rating"])
    certificate_rows = load_table_rows(workbook["certificate"])
    stats_rows = load_table_rows(workbook["certificate_statistics"])

    rating_map = build_rating_map(rating_rows)
    certificates = build_certificate_rows(certificate_rows)
    stats_entries = build_statistics(stats_rows)

    stats_by_cert: Dict[str, List[StatisticEntry]] = defaultdict(list)
    for entry in stats_entries:
        stats_by_cert[entry.cert_id].append(entry)

    documents: List[Dict[str, Any]] = []
    for cert in certificates.values():
        documents.append(build_profile_document(cert, rating_map))
        stat_docs = build_statistics_documents(cert, stats_by_cert.get(cert.cert_id, []))
        documents.extend(stat_docs)
    return documents


def main() -> None:
    parser = argparse.ArgumentParser(description="Build RAG documents from the SkillBridge data workbook.")
    parser.add_argument("--input", type=Path, default=Path("data/data.xlsx"), help="Path to the Excel workbook.")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/rag/documents.jsonl"),
        help="Destination path for the generated JSONL file.",
    )
    args = parser.parse_args()

    if not args.input.exists():
        raise SystemExit(f"Input file not found: {args.input}")

    documents = generate_documents(args.input)
    args.output.parent.mkdir(parents=True, exist_ok=True)

    with args.output.open("w", encoding="utf-8") as fp:
        for doc in documents:
            fp.write(json.dumps(doc, ensure_ascii=False))
            fp.write("\n")

    print(f"Generated {len(documents)} documents at {args.output}")


if __name__ == "__main__":
    main()
